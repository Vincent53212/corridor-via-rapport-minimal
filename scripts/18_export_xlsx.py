#!/usr/bin/env python3
"""Étape 18 — Export des livrables CSV en .xlsx multi-colonnes filtrables.

Problème résolu : les CSV (séparateur « ; ») s'ouvrent parfois en UNE seule
colonne dans Excel selon la locale. On produit donc, À CÔTÉ de chaque CSV, un
vrai .xlsx avec colonnes séparées, en-tête figé en gras et **filtre par colonne**
(AutoFilter). Les valeurs numériques sont converties en nombres (tri correct).

Étage de pipeline : relancer ce script régénère tous les .xlsx à partir des CSV
courants (donc ils se mettent à jour quand les scripts amont changent).

Dépendances : stdlib `csv` + `openpyxl` (indépendant du venv géo).
Sortie : livrables/<nom>.xlsx pour chaque livrables/<nom>.csv.
"""
from __future__ import annotations
import csv
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import DELIVERABLES

# Fichiers à double en-tête FR/EN (lignes 1=FR, 2=EN) → fusionnées « FR / EN ».
DUAL_HEADER = {"segments_courbature.csv"}

NUM_RE = re.compile(r"^-?\d+(\.\d+)?$")
HEADER_FILL = PatternFill("solid", fgColor="003366")
HEADER_FONT = Font(bold=True, color="FFFFFF")
MAX_WIDTH = 46


def coerce(v: str):
    """Convertit une cellule texte en nombre si possible (tri numérique Excel)."""
    if v is None or v == "":
        return None
    if NUM_RE.match(v):
        f = float(v)
        return int(f) if ("." not in v and -2**53 < f < 2**53) else f
    return v


def read_table(path: Path) -> tuple[list[str], list[list]]:
    """Retourne (header, data_rows). Gère le double en-tête et le séparateur « ; »."""
    with path.open(encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f, delimiter=";"))
    if not rows:
        return [], []
    if path.name in DUAL_HEADER and len(rows) >= 2:
        header = [f"{fr} / {en}" for fr, en in zip(rows[0], rows[1])]
        data = rows[2:]
    else:
        header = rows[0]
        data = rows[1:]
    return header, data


def write_xlsx(header: list[str], data: list[list], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = out_path.stem[:31] or "Données"

    ws.append(header)
    for c, _ in enumerate(header, 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=False)

    for row in data:
        ws.append([coerce(v) for v in row])

    n_rows = len(data) + 1
    n_cols = len(header)
    if n_cols:
        ref = f"A1:{get_column_letter(n_cols)}{n_rows}"
        ws.auto_filter.ref = ref           # filtre par colonne
    ws.freeze_panes = "A2"                  # en-tête figé

    # Largeurs de colonnes : max du contenu, plafonné.
    for c in range(1, n_cols + 1):
        letter = get_column_letter(c)
        longest = len(str(header[c - 1]))
        for row in data:
            if c - 1 < len(row):
                longest = max(longest, len(str(row[c - 1])))
        ws.column_dimensions[letter].width = min(max(longest + 2, 8), MAX_WIDTH)

    wb.save(out_path)


def main() -> None:
    csvs = sorted(DELIVERABLES.glob("*.csv"))
    if not csvs:
        sys.exit(f"Aucun CSV dans {DELIVERABLES}")
    print("=== Étape 18 — Export CSV → .xlsx (multi-colonnes + filtres) ===")
    n_ok = 0
    for path in csvs:
        header, data = read_table(path)
        if not header:
            print(f"  (ignoré, vide) {path.name}")
            continue
        # Fichiers « message » (1 colonne, pas une table) : on passe.
        if len(header) <= 1 and len(data) <= 1:
            print(f"  (ignoré, non tabulaire) {path.name}")
            continue
        out = path.with_suffix(".xlsx")
        write_xlsx(header, data, out)
        print(f"  {out.name}  ({len(data)} lignes × {len(header)} colonnes)")
        n_ok += 1
    print(f"\n{n_ok} fichier(s) .xlsx écrit(s) dans {DELIVERABLES.name}/")


if __name__ == "__main__":
    main()
